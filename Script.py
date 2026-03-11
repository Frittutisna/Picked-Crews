import  argparse
import  random
import  re
import  sys
import  traceback
from    openpyxl    import  load_workbook
from    pathlib     import  Path

# Configurations
SPORTS_MODES    = ["MLB", "NBA", "NFL", "F1"]
DEFAULT_MODES   = 50
DEFAULT_SPORTS  = 0

# Script
def get_xlsx_file():
    '''Finds the first XLSX file in the current directory'''
    files = list(Path('.').glob('*.xlsx'))
    if not files:
        print("[!] Error: No XLSX file found in the current folder")
        sys.exit(1)
    return files[0]

def read_mode_data(xlsx_path: Path, sheet_name: str = "Descriptions") -> list[dict]:
    '''Reads names, row numbers, and formatted player count values'''
    wb      = load_workbook(xlsx_path, read_only = True, data_only = True)
    ws      = wb[sheet_name] if sheet_name in wb.sheetnames else wb.worksheets[0]
    data    = []
    
    for row in ws.iter_rows(min_row = 2, max_col = 3, values_only = False): 
        name_val = row[0].value
        if name_val:
            raw_c   = str(row[2].value or "").strip()
            if not raw_c: continue
            clean_c = raw_c.lower().split('v')[0] if 'v' in raw_c.lower() else raw_c            
            if clean_c.strip(): data.append({
                "name"  : str(name_val).strip(),
                "row"   : row[0].row,
                "val_c" : int(clean_c.strip())
            })
    return data

def parse_setup():
    '''Parses Setup.txt into a structured dictionary'''
    content             = Path("Setup.txt").read_text(encoding = "utf-8")
    data                = {}
    data["size"]        = int(re.search(r"Size:\s*(\d+)", content).group(1))
    data["protected"]   = [int(x) for x in re.search(r"Protected:\s*([\d,\s]+)", content).group(1).replace(',', ' ').split()]
    data["banned"]      = [tuple(int(x) for x in m.split(',')) for m in re.findall(r"\((\d+(?:,\s*\d+)*)\)", content.split("Banned:")[1].split("Picked:")[0])]
    data["picked"]      = [tuple(int(x) for x in m.split(',')) for m in re.findall(r"\((\d+(?:,\s*\d+)*)\)", content.split("Picked:")[1])]
    return data

def parse_rolls():
    '''Parses existing Rolls.txt back into dictionaries'''
    lines   = Path("Rolls.txt").read_text(encoding = "utf-8").splitlines()
    rolled  = []
    for line in lines:
        match = re.match(r"(\d+)\.\s+(.*)", line)
        if match:
            idx         = int(match.group(1))
            full_name   = match.group(2).strip()
            count_match = re.search(r"(\d+)v\d+", full_name)
            val_c       = int(count_match.group(1)) if count_match else 1
            rolled.append({
                "list_idx"  : idx,
                "name"      : full_name, 
                "val_c"     : val_c
            })
    return rolled

def validate_setup(setup, rolled_map):
    '''Validates constraints against Setup.txt'''
    if len(setup["protected"]) != 2: 
        raise ValueError(f"Team A and B must each protect exactly 1 mode, found {len(setup['protected'])}")

    for i in range(2):
        team_label = "Team A" if i == 0 else "Team B"
        opp_label  = "Team B" if i == 0 else "Team A"
        opp_idx    = 1 - i

        opp_protected_idx = setup["protected"][opp_idx]
        if opp_protected_idx in setup["banned"][i]:
            raise ValueError(f"{team_label} banned the protected {rolled_map[opp_protected_idx]['name']}")

        banned_by_opponent  = set(setup["banned"][opp_idx])
        team_picks          = set(setup["picked"][i])
        clashes             = team_picks & banned_by_opponent
        if clashes:
            raise ValueError(f"{team_label} picked mode(s) banned by {opp_label}: {', '.join([rolled_map[c]['name'] for c in clashes])}")

        total_banned = sum(rolled_map[idx]["val_c"] for idx in setup["banned"][i])
        if total_banned != setup["size"]:
            raise ValueError(f"Modes banned by {team_label} for {opp_label} total {total_banned} players, expected {setup['size']}")
        
        team_protected_idx = setup["protected"][i]
        if team_protected_idx in setup["picked"][i]:
            raise ValueError(f"{team_label} picked the already-protected {rolled_map[team_protected_idx]['name']}")

        total_picked = rolled_map[team_protected_idx]["val_c"] + sum(rolled_map[idx]["val_c"] for idx in setup["picked"][i])
        if total_picked != setup["size"]:
            raise ValueError(f"Modes protected/picked by {team_label} total {total_picked} players, expected {setup['size']}")

def format_mode_name(name, val_c):
    '''Appends nvn suffix if not already present in the name'''
    suffix = f"{val_c}v{val_c}"
    if suffix in name.lower(): return name
    return f"{name} {suffix}"

def get_core_name(name):
    '''Strips Random/Watched/Mixed/Unwatched prefix and (BR/(D)NGMC) nvn suffixes'''
    core = re.sub(r'^(Random|Watched|Mixed|Unwatched)\s+',          '', name, flags = re.IGNORECASE)
    core = re.sub(r'\s+(\d+v\d+|BR\s+\d+v\d+|(D?NGMC)\s+\d+v\d+)$', '', core, flags = re.IGNORECASE)
    return core.strip()

def generate_rolls(all_data, args):
    '''Generates Rolls.txt based on arguments, core mode constrains, and minimum requirements'''
    n           = max(50, min(100, args.modes))
    min_req     = 5 * ((n // 4) // 5)
    attempts    = 0
    selected    = []

    while attempts < 10000:
        sample_pool         = random.sample(all_data, len(all_data))
        current_selection   = []
        selected_cores      = set()

        for item in sample_pool:
            core = get_core_name(item["name"])
            if core not in selected_cores:
                current_selection.append(item)
                selected_cores.add(core)
            if len(current_selection) == n: break

        watched_count   = sum(1 for d in current_selection if "watched"     in d["name"].lower() and "unwatched" not in d["name"].lower())
        random_count    = sum(1 for d in current_selection if "random"      in d["name"].lower())
        spotlight_count = sum(1 for d in current_selection if "spotlight"   in d["name"].lower())
        sports_count    = sum(1 for d in current_selection if any(s.lower() in d["name"].lower() for s in SPORTS_MODES))

        v1_count        = sum(1 for d in current_selection if d["val_c"] == 1)
        v2_count        = sum(1 for d in current_selection if d["val_c"] == 2)
        v3_count        = sum(1 for d in current_selection if d["val_c"] == 3)
        v4_count        = sum(1 for d in current_selection if d["val_c"] == 4)

        if (
            watched_count   >= min_req      and 
            random_count    >= min_req      and 
            spotlight_count >= min_req      and
            sports_count    >= args.sports  and
            v1_count        >= min_req      and
            v2_count        >= min_req      and
            v3_count        >= min_req      and
            v4_count        >= min_req
        ):    
            selected = current_selection
            break

        attempts += 1

    if not selected: raise ValueError(f"Could not find a valid list of Rolled Modes, try again")
    selected.sort(key = lambda x: x["name"].lower())
    output_lines    = [f"{i}. {format_mode_name(item['name'], item['val_c'])}" for i, item in enumerate(selected, 1)]
    output_text     = "Rolled Modes: \n" + "\n".join(output_lines)
    Path("Rolls.txt").write_text(output_text, encoding = "utf-8")
    print(output_text)
    print(f"[✓] Success: Generated Rolls.txt, copy-paste it in #tour-information")

def generate_results(setup, rolled_list):
    '''Generates Results.txt based on Setup.txt and Rolls.txt'''
    rolled_map = {item["list_idx"]: item for item in rolled_list}
    validate_setup(setup, rolled_map)

    rounds = []
    for i in range(2):
        team_indices    = [setup["protected"][i]] + list(setup["picked"][i])
        team_modes      = sorted([rolled_map[idx]['name'] for idx in team_indices], key = lambda x: x.lower())
        rounds.append([f"{m}: " for m in team_modes])

    used_indices    = set(setup["protected"]) | {i for t in setup["banned"] for i in t} | {i for t in setup["picked"] for i in t}
    pool            = [d for d in rolled_list if d["list_idx"] not in used_indices]
    round_3_final   = []
    attempts        = 0
    while attempts < 1000:
        sample          = []
        current_p       = 0
        spotlight_count = 0
        shuffled_pool   = random.sample(pool, len(pool))
        
        for item in shuffled_pool:
            is_spotlight = "spotlight" in item["name"].lower()
            if current_p + item["val_c"] <= setup["size"]:
                if is_spotlight and spotlight_count >= 1    : continue
                sample.append(item)
                current_p += item["val_c"]
                if is_spotlight                             : spotlight_count += 1
        
        watched_modes = [d for d in sample if "watched" in d["name"].lower()]
        random_modes  = [d for d in sample if "random"  in d["name"].lower()]
        if current_p == setup["size"] and abs(len(watched_modes) - len(random_modes)) <= 1:
            sample.sort(key = lambda x: x["name"].lower())
            round_3_final = [f"{d['name']}: " for d in sample]
            break
        attempts += 1
    else: raise ValueError("Could not find a valid Round 3 combination matching constraints")

    output =    f"Round 1: \n" + "\n".join(f"{i+1}. {line} " for i, line in enumerate(rounds[0]))       + "\n\n"
    output +=   f"Round 2: \n" + "\n".join(f"{i+1}. {line} " for i, line in enumerate(rounds[1]))       + "\n\n"
    output +=   f"Round 3: \n" + "\n".join(f"{i+1}. {line} " for i, line in enumerate(round_3_final))
    
    Path("Results.txt").write_text(output, encoding = "utf-8")
    print(output)
    print("[✓] Success: Generated Results.txt, copy-paste it in #tour-information")

def main():
    parser = argparse.ArgumentParser(description = "[?] Roll modes for Picked Crews")
    parser.add_argument("--modes",  type = int, default = DEFAULT_MODES,  help = "[?] Number of modes to roll")
    parser.add_argument("--sports", type = int, default = DEFAULT_SPORTS, help = "[?] Minimum number of Sports Modes to roll")
    
    args        = parser.parse_args()
    xlsx_path   = get_xlsx_file()
    
    if Path("Setup.txt").exists() and Path("Rolls.txt").exists():
        print("[.] Setup.txt found, generating Results.txt")
        try:
            rolled_list = parse_rolls()
            setup_data  = parse_setup()
            generate_results(setup_data, rolled_list)
            return
        except Exception as e:
            print(f"[!] Error: {e}")
            sys.exit(1)
    
    else:
        print(f"[.] Setup.txt not found, generating Rolls.txt")
        try:
            all_data = read_mode_data(xlsx_path)
            generate_rolls(all_data, args)
        except Exception as e:
            traceback.print_exc()
            sys.exit(1)

if __name__ == "__main__": main()